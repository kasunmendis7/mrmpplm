from openpyxl import Workbook
import os

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def adjust_column_widths(sheet, width=0):
    # Loop through all columns

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        if width == 0:
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Adding some padding
            sheet.column_dimensions[column_letter].width = adjusted_width
        else:
            sheet.column_dimensions[column_letter].width = width


def align_center(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


optimizations = ['Vanilla', 'XBZRLE', 'Compress', 'HBFDP_SHA1', 'HBFDP_MD5', 'HBFDP_Murmur3', 'Dtrack', 'Compress_Dtrack', 'XBZRLE_Dtrack']
# 'XBZRLE_HBFDP_SHA1', 'XBZRLE_HBFDP_MD5', 'XBZRLE_HBFDP_Murmur3', 'Compress_HBFDP_SHA1', 'Compress_HBFDP_MD5', 'Compress_HBFDP_Murmur3',
workloads = ['Memcached', 'Sysbench', 'Oltp']
rams = [1, 2, 4, 8, 12, 16]
cells = [3, 5]
average = {
    "Iterations": 'B',
    "Setup-Time": 'C',
    "Downtime": 'D',
    "Total-Time": 'E',
    "Ram-Total": 'F',
    "Ram-Postcopy-Requests": 'G',
    "Ram-Dirty-Sync-Count": 'H',
    "Ram-Remaining": 'I',
    "Ram-Mbps": 'J',
    "Ram-Transferred": 'K',
    "Ram-Duplicate": 'L',
    "Ram-Dirty-Pages-Rate": 'M',
    "Ram-Skipped": 'N',
    "Ram-Normal-Bytes": 'O',
    "Ram-Normal": 'P',
    "Total-Fake-Dirty-Pages": 'Q',
    "Total-Transferred-Page-Count": 'R'
}

folder = '.\\data\\Analysis'

for workload in workloads:

    wb = Workbook()
    wb_arr = []
    wb_count = 0
    wSummary = wb.active
    wSummary.title = "Summary"

    for optimization in optimizations:

        wSummary.append(["Optimization", f"{optimization}"])
        wSummary.append(["Ram", "Total-Time", "Total-Fake-Dirty-Pages", "Total-Transferred-Page-Count"])

        cells = [3, 5]

        wb_arr.append(wb.create_sheet(title=f"{optimization}"))

        for ram in rams:
            ram_mb = str(ram * 1024)

            if not os.path.exists(f".\\data\\{optimization}\\{workload}\\{ram_mb}.csv"):
                continue

            exp_count = 1

            with open(f".\\data\\{optimization}\\{workload}\\{ram_mb}.csv", 'r') as f:

                wb_arr[wb_count].append(["RAM", int(ram_mb)])

                lines = f.readlines()
                for line in lines:
                    line = line.strip("\n").split(",")
                    if "No" in line[1] or "Outlier" in line[1]:
                        del line[1]

                        if "Experiment" not in line[0]:
                            line[0] = exp_count
                            exp_count += 1
                            for i in range(1, 18):
                                if i == 9:
                                    line[i] = float(line[i])
                                else:
                                    line[i] = int(line[i])

                        wb_arr[wb_count].append(line)

                arr = ["Average"]
                count = 1
                for key in average.keys():
                    arr.append(f'=AVERAGE({average[key]}{cells[0]}:{average[key]}{cells[1]})')
                wb_arr[wb_count].append(arr)
                wb_arr[wb_count].append([""])

                d4 = f"='{optimization}'!E{cells[1] + 1}"
                d16 = f"='{optimization}'!Q{cells[1] + 1}"
                d17 = f"='{optimization}'!R{cells[1] + 1}"
                wSummary.append([int(ram_mb), d4, d16, d17])

            cells[0] += 7
            cells[1] += 7

        wSummary.append([""])
        adjust_column_widths(wb_arr[wb_count], 0)
        align_center(wb_arr[wb_count])
        wb_count += 1

    adjust_column_widths(wSummary, 25)
    align_center(wSummary)
    wb.save(f"{folder}\\{workload}_Analysis.xlsx")
